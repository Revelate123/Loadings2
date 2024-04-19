import PySimpleGUI as sg
import os
def KeyCheck(Dictionary, string):
    try:
        output = Dictionary[string]
        if output == 'True':
            output = 1
        elif output == 'False':
            output = 0
        if output == '':
            output = 0
    except KeyError:
        output = 0
    return output
def KeyCheck1(Dictionary, string):
    try:
        output = Dictionary[string]
        if output == 'True':
            output = 1
        elif output == 'False':
            output = 0
        if output == '':
            output = 1
    except KeyError:
        output = 1
    return output


def Cantilever_Layout(variable):
    Layout = [[sg.Column([
        [sg.Text('Cantilever Length')],
    ]), sg.Column([
        [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'CLength'), enable_events=True, key='CLength'),sg.Button('Update', key = 'CLength_Update')],
    ])],
        [sg.Text('Number of UDL\'s'),
         sg.Input(default_text=int(float(KeyCheck(variable, 'CUDLS'))), key='CUDLS', enable_events=True, size=(5, 1))]
    ]
    try:
        CUDLS = int(float(KeyCheck(variable, 'CUDLS')))
    except:
        CUDLS = 1
    for i in range(CUDLS):
        i = str(i) + 'C'
        Layout += [[sg.Column([
            [sg.Text('Load Case')],
            [sg.Text('G1')],
            [sg.Text('G2')],
            [sg.Text('Q')],
            [sg.Text('Wup')],
            [sg.Text('Wdown')],
            [sg.Text('WOoP')],
        ]),
            sg.Column([
                [sg.Text('Magnitude KPa')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1M' + str(i)), enable_events=True,
                          key='G1M' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2M' + str(i)), enable_events=True,
                          key='G2M' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QM' + str(i)), enable_events=True,
                          key='QM' + str(i))],
                [sg.Text()],
                [sg.Text()],
                [sg.Text()]
                # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WupM'+str(i)),enable_events=True,key='WupM'+str(i))],
                # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WdownM'+str(i)),enable_events=True,key='WdownM'+str(i))],
                # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WOoPM'+str(i)),enable_events=True,key='WOoPM'+str(i))]
            ]),
            sg.Column([
                [sg.Text('Tributary Width m')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1T' + str(i)), enable_events=True,
                          key='G1T' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2T' + str(i)), enable_events=True,
                          key='G2T' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QT' + str(i)), enable_events=True,
                          key='QT' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupT' + str(i)), enable_events=True,
                          key='WupT' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownT' + str(i)), enable_events=True,
                          key='WdownT' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPT' + str(i)), enable_events=True,
                          key='WOoPT' + str(i))]
            ]),
            sg.Column([
                [sg.Text('Start m')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1a' + str(i)), enable_events=True,
                          key='G1a' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2a' + str(i)), enable_events=True,
                          key='G2a' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Qa' + str(i)), enable_events=True,
                          key='Qa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wupa' + str(i)), enable_events=True,
                          key='Wupa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wdowna' + str(i)), enable_events=True,
                          key='Wdowna' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPa' + str(i)), enable_events=True,
                          key='WOoPa' + str(i))],
            ]),
            sg.Column([
                [sg.Text('Stop m')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1b' + str(i)), enable_events=True,
                          key='G1b' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2b' + str(i)), enable_events=True,
                          key='G2b' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Qb' + str(i)), enable_events=True,
                          key='Qb' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wupb' + str(i)), enable_events=True,
                          key='Wupb' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wdownb' + str(i)), enable_events=True,
                          key='Wdownb' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPb' + str(i)), enable_events=True,
                          key='WOoPb' + str(i))],
            ])]]
    Layout += [[sg.Text('New point Loads'),
                sg.Input(key='CPoint_loads', default_text=int(float(KeyCheck(variable, 'CPoint_loads'))),
                         enable_events=True, size=(5, 1))]]
    try:
        CPoint_loads = int(float(KeyCheck(variable, 'CPoint_loads')))
    except:
        CPoint_loads = 1
    for i in range(CPoint_loads):
        i = str(i) + 'C'
        Layout += [[sg.Column([
            [sg.Text('Load Case')],
            [sg.Text('G1')],
            [sg.Text('G2')],
            [sg.Text('Q')],
            [sg.Text('Wup')],
            [sg.Text('Wdown')],
            [sg.Text('WOoP')],
        ]),
            sg.Column([
                [sg.Text('Magnitude KN')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1P' + str(i)), enable_events=True,
                          key='G1P' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2P' + str(i)), enable_events=True,
                          key='G2P' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QP' + str(i)), enable_events=True,
                          key='QP' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupP' + str(i)), enable_events=True,
                          key='WupP' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownP' + str(i)), enable_events=True,
                          key='WdownP' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPP' + str(i)), enable_events=True,
                          key='WOoPP' + str(i))]
            ]),
            sg.Column([
                [sg.Text('Location')],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1Pa' + str(i)), enable_events=True,
                          key='G1Pa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2Pa' + str(i)), enable_events=True,
                          key='G2Pa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QPa' + str(i)), enable_events=True,
                          key='QPa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupPa' + str(i)), enable_events=True,
                          key='WupPa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownPa' + str(i)), enable_events=True,
                          key='WdownPa' + str(i))],
                [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPPa' + str(i)), enable_events=True,
                          key='WOoPPa' + str(i))]
            ])]]

    Layout += [[sg.Text('Existing point loads'),
                sg.Input(key='CEx_Point_loads', default_text=int(float(KeyCheck(variable, 'CEx_Point_loads'))),
                         enable_events=True,
                         size=(5, 1))]]
    try:
        directory = KeyCheck(variable, 'Project')
    except:
        directory = os.fsencode(os.getcwd())
    existing = []
    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        if filename.endswith(".pkl"):
            filename = os.path.basename(filename)
            filename = os.path.splitext(filename)[0]
            if filename != KeyCheck(variable, 'Name'):
                existing += [filename]
            continue
        else:
            continue
    try:
        CEx_Point_loads = int(float(KeyCheck(variable, 'CEx_Point_loads')))
    except:
        CEx_Point_loads = 1
    for i in range(CEx_Point_loads):
        i = str(i) + 'C'
        try:
            Layout += [[sg.Combo(existing, size=(10, 1), key='Ex' + str(i),
                                 default_value=KeyCheck(variable, 'Ex' + str(i)), enable_events=True),
                        sg.Combo(['Left', 'Right'], size=(5, 1), key='L/R' + str(i),
                                 default_value=KeyCheck(variable, 'L/R' + str(i)), enable_events=True),
                        sg.Input(default_text=KeyCheck(variable, 'Exa' + str(i)), key='Exa' + str(i), size=(5, 1),
                                 enable_events=True)]]
        except:
            Ex_Point_loads = 0
    return Layout


def Layouts(variable):
    from openpyxl import load_workbook

    wb = load_workbook(filename='Steel Design Calculator.xlsx')
    # SectionType = wb.sheetnames
    SectionSize = []
    # print(sheet_ranges['A6'].value)
    SectionType = ['Universal_Beam', 'Universal_Column', 'PFC', 'RHS', 'SHS', 'CHS', 'MGP 10', 'MGP 12']
    for j in wb['Universal_Beam']:
        SectionSize.append(j[0].value)
    SectionSize = list(filter(lambda item: item is not None, SectionSize))

    for j1 in range(4):
        SectionSize.pop(0)

    Layout = [[sg.Column([
        [sg.Text('Current Project'),
         sg.Input(key='Project', default_text=KeyCheck(variable, 'Project'), enable_events=True),
         sg.FolderBrowse('Change Project', enable_events=True, key='Project')],
        [sg.Column([
            [sg.Text('Name')],
            [sg.Text('Length')],
            [sg.Text('E')],
            [sg.Text('Choose Section Type:')],
            [sg.Text('Choose Section Size:')],
            [sg.Text('Seasoned')],
            [sg.Text('Duration of Load')],
            [sg.Text('Initial Moisture content:')],
            [sg.Text('Ix')],
            [sg.Text('Iy')]]), sg.Column([
            [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'Name'), enable_events=True, key='Name'),
             sg.Button('Reload Existing', key='Reload')],
            [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'Length'), enable_events=True, key='Length'),
             sg.Button('Update', key='Length_Update'),
             sg.Checkbox('Cantilever', default=KeyCheck(variable, 'Cantilever'), key='Cantilever', enable_events=True)],
            [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'E'), enable_events=True, key='E'),
             sg.Checkbox('Override for E', default=KeyCheck(variable, 'E_check'), key='E_check', enable_events=True)],
            [sg.Combo(SectionType, key='SectionType', enable_events=True,
                      default_value=KeyCheck(variable, 'SectionType'),
                      size=(30, 1)),
             sg.Checkbox('Self Weight', default=KeyCheck(variable, 'SelfWeight'), key='SelfWeight',
                         enable_events=True)],
            # SectionType[0]
            [sg.Combo(SectionSize, key='SectionSize', enable_events=True,
                      default_value=KeyCheck(variable, 'SectionSize'),
                      size=(30, 1))],  # SectionSize[0]
            [sg.Combo(['Seasoned', 'Unseasoned'], default_value='Seasoned', key='Seasoned')],
            [sg.Combo(['5 seconds', '5 minutes', '5 hours', '5 days', '5 months', '50+ years'],
                      default_value='5 months', key='load_duration')],
            [sg.Combo(['<15%', '>25%'], key='j2', default_value=KeyCheck(variable, 'j2'), size=(10, 1))],
            [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'Ix'), enable_events=True, key='Ix'),
             sg.Checkbox('Override for Ix', default=KeyCheck(variable, 'Ix_check'), key='Ix_check',
                         enable_events=True)],
            [sg.Input(size=(10, 1), default_text=KeyCheck(variable, 'Iy'), enable_events=True, key='Iy'),
             sg.Checkbox('Override for Iy', default=KeyCheck(variable, 'Iy_check'), key='Iy_check',
                         enable_events=True)]])]])]]
    Layout += steel_calculator(variable)
    Layout += [[sg.Column([
        [sg.Text('Base Wind pressure (ULS):'),
         sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'P'), enable_events=True, key='P')],
        [sg.Text('Vertical Cp,e:'),
         sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Cpe_V'), enable_events=True, key='Cpe_V')],
        [sg.Text('Vertical Cp,i:'),
         sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Cpi_V'), enable_events=True, key='Cpi_V')],
        [sg.Text('Horizontal Cp,e:'),
         sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Cpe_H'), enable_events=True, key='Cpe_H')],
        [sg.Text('Horizontal Cp,i:'),
         sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Cpi_H'), enable_events=True, key='Cpi_H')],
        [sg.Text('Number of UDL\'s'),
         sg.Input(default_text=int(float(KeyCheck(variable, 'UDLS'))), key='UDLS', enable_events=True, size=(5, 1))]
    ])]]
    try:
        UDLS = int(float(KeyCheck(variable, 'UDLS')))
    except:
        UDLS = 1
    for i in range(UDLS):
        Layout += [[sg.Input(default_text=KeyCheck(variable, 'Load_case' + str(i)), enable_events=True,
                             key='Load_case' + str(i))], [sg.Column([
            [sg.Text('Load Case')],
            [sg.Text('G1')],
            [sg.Text('G2')],
            [sg.Text('Q')],
            [sg.Text('Wup')],
            [sg.Text('Wdown')],
            [sg.Text('WOoP')],
        ]),
                       sg.Column([
                           [sg.Text('Magnitude KPa')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1M' + str(i)), enable_events=True,
                                     key='G1M' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2M' + str(i)), enable_events=True,
                                     key='G2M' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QM' + str(i)), enable_events=True,
                                     key='QM' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupM' + str(i)), enable_events=True,
                                     key='WupM' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownM' + str(i)),
                                     enable_events=True, key='WdownM' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPM' + str(i)), enable_events=True,
                                     key='WOoPM' + str(i))]
                           # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WupM'+str(i)),enable_events=True,key='WupM'+str(i))],
                           # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WdownM'+str(i)),enable_events=True,key='WdownM'+str(i))],
                           # [sg.Input(size=(5,1),default_text=KeyCheck(variable,'WOoPM'+str(i)),enable_events=True,key='WOoPM'+str(i))]
                       ]),
                       sg.Column([
                           [sg.Text('Tributary Width m')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1T' + str(i)), enable_events=True,
                                     key='G1T' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2T' + str(i)), enable_events=True,
                                     key='G2T' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QT' + str(i)), enable_events=True,
                                     key='QT' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupT' + str(i)), enable_events=True,
                                     key='WupT' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownT' + str(i)),
                                     enable_events=True, key='WdownT' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPT' + str(i)), enable_events=True,
                                     key='WOoPT' + str(i))]
                       ]),
                       sg.Column([
                           [sg.Text('Start m')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1a' + str(i)), enable_events=True,
                                     key='G1a' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2a' + str(i)), enable_events=True,
                                     key='G2a' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Qa' + str(i)), enable_events=True,
                                     key='Qa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wupa' + str(i)), enable_events=True,
                                     key='Wupa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wdowna' + str(i)),
                                     enable_events=True, key='Wdowna' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPa' + str(i)), enable_events=True,
                                     key='WOoPa' + str(i))],
                       ]),
                       sg.Column([
                           [sg.Text('Stop m')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1b' + str(i)), enable_events=True,
                                     key='G1b' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2b' + str(i)), enable_events=True,
                                     key='G2b' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Qb' + str(i)), enable_events=True,
                                     key='Qb' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wupb' + str(i)), enable_events=True,
                                     key='Wupb' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'Wdownb' + str(i)),
                                     enable_events=True, key='Wdownb' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPb' + str(i)), enable_events=True,
                                     key='WOoPb' + str(i))],
                       ])]]
    Layout += [[sg.Text('New point Loads'),
                sg.Input(key='Point_loads', default_text=int(float(KeyCheck(variable, 'Point_loads'))),
                         enable_events=True, size=(5, 1))]]
    try:
        Point_loads = int(float(KeyCheck(variable, 'Point_loads')))
    except:
        Point_loads = 1
    for i in range(Point_loads):
        Layout += [[sg.Input(default_text=KeyCheck(variable, 'Load_caseP' + str(i)), enable_events=True,
                             key='Load_caseP' + str(i))], [sg.Column([
            [sg.Text('Load Case')],
            [sg.Text('G1')],
            [sg.Text('G2')],
            [sg.Text('Q')],
            [sg.Text('Wup')],
            [sg.Text('Wdown')],
            [sg.Text('WOoP')],
        ]),
                       sg.Column([
                           [sg.Text('Magnitude KN')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1P' + str(i)), enable_events=True,
                                     key='G1P' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2P' + str(i)), enable_events=True,
                                     key='G2P' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QP' + str(i)), enable_events=True,
                                     key='QP' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupP' + str(i)), enable_events=True,
                                     key='WupP' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownP' + str(i)),
                                     enable_events=True, key='WdownP' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPP' + str(i)), enable_events=True,
                                     key='WOoPP' + str(i))]
                       ]),
                       sg.Column([
                           [sg.Text('Location')],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G1Pa' + str(i)), enable_events=True,
                                     key='G1Pa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'G2Pa' + str(i)), enable_events=True,
                                     key='G2Pa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'QPa' + str(i)), enable_events=True,
                                     key='QPa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WupPa' + str(i)), enable_events=True,
                                     key='WupPa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WdownPa' + str(i)),
                                     enable_events=True, key='WdownPa' + str(i))],
                           [sg.Input(size=(5, 1), default_text=KeyCheck(variable, 'WOoPPa' + str(i)),
                                     enable_events=True, key='WOoPPa' + str(i))]
                       ])]]

    Layout += [[sg.Text('Existing point loads'),
                sg.Input(key='Ex_Point_loads', default_text=int(float(KeyCheck(variable, 'Ex_Point_loads'))),
                         enable_events=True,
                         size=(5, 1))]]
    try:
        directory = KeyCheck(variable, 'Project')
        existing = []
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            if filename.endswith(".pkl"):
                filename = os.path.basename(filename)
                filename = os.path.splitext(filename)[0]
                if filename != KeyCheck(variable, 'Name'):
                    existing += [filename]
                continue
            else:
                continue
    except:
        directory = os.fsencode(os.getcwd())
        existing = []
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            if filename.endswith(".pkl"):
                filename = os.path.basename(filename)
                filename = os.path.splitext(filename)[0]
                if filename != KeyCheck(variable, 'Name'):
                    existing += [filename]
                continue
            else:
                continue
    try:
        Ex_Point_loads = int(float(KeyCheck(variable, 'Ex_Point_loads')))
    except:
        Ex_Point_loads = 1
    for i in range(Ex_Point_loads):
        try:
            Layout += [[sg.Combo(existing, size=(10, 1), key='Ex' + str(i),
                                 default_value=KeyCheck(variable, 'Ex' + str(i)), enable_events=True),
                        sg.Combo(['Left', 'Right'], size=(5, 1), key='L/R' + str(i),
                                 default_value=KeyCheck(variable, 'L/R' + str(i)), enable_events=True),
                        sg.Input(default_text=KeyCheck(variable, 'Exa' + str(i)), key='Exa' + str(i), size=(5, 1),
                                 enable_events=True)]]
        except:
            Ex_Point_loads = 0

    Layout += [[sg.Column([
        [sg.Button('Calculate', key='Calculate')],
        [sg.Text('Number of Load Cases:'),
         sg.Input(key='Load_Cases', default_text=int(float(KeyCheck(variable, 'Load_Cases'))), size=(5, 1)),
         sg.Button('Default Load Cases', key='Default_Load_Cases')],
    ])]]
    Layout += [[sg.Text('', key='P/F_Check')]]
    try:
        Load_Cases = int(float(KeyCheck(variable, 'Load_Cases')))
    except:
        Load_Cases = 1

    Canvas = []
    for i in range(Load_Cases):
        # try:
        COL1 = [[sg.Text('Load Case')], [
            sg.Input(key='Load_Case' + str(i), default_text=KeyCheck(variable, 'Load_Case' + str(i)), size=(20, 1),
                     enable_events=True)]]
        COL2 = [[sg.Text('G1')],
                [sg.Input(key='Load_CaseG1' + str(i), default_text=KeyCheck(variable, 'Load_CaseG1' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL3 = [[sg.Text('G2')],
                [sg.Input(key='Load_CaseG2' + str(i), default_text=KeyCheck(variable, 'Load_CaseG2' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL4 = [[sg.Text('Q')],
                [sg.Input(key='Load_CaseQ' + str(i), default_text=KeyCheck(variable, 'Load_CaseQ' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL5 = [[sg.Text('Wup')],
                [sg.Input(key='Load_CaseWup' + str(i), default_text=KeyCheck(variable, 'Load_CaseWup' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL6 = [[sg.Text('Wdown')],
                [sg.Input(key='Load_CaseWdown' + str(i), default_text=KeyCheck(variable, 'Load_CaseWdown' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL7 = [[sg.Text('WOoP')],
                [sg.Input(key='Load_CaseWOoP' + str(i), default_text=KeyCheck(variable, 'Load_CaseWOoP' + str(i)),
                          size=(5, 1), enable_events=True)]]
        COL8 = [[sg.Text('Moment')],
                [sg.Checkbox('Moment', key='Load_CaseM' + str(i),
                             default=bool(float(KeyCheck(variable, 'Load_CaseM' + str(i)))), enable_events=True)]]
        COL9 = [[sg.Text('Shear')],
                [sg.Checkbox('Shear', key='Load_CaseS' + str(i),
                             default=bool(float(KeyCheck(variable, 'Load_CaseS' + str(i)))), enable_events=True)]]
        COL10 = [[sg.Text('Deflection')],
                 [sg.Checkbox('Deflection', key='Load_CaseD' + str(i),
                              default=bool(float(KeyCheck(variable, 'Load_CaseD' + str(i)))), enable_events=True)]]
        COL11 = [[sg.Text('Deflection Limit L/')], [
            sg.Input(default_text=KeyCheck1(variable, 'Load_CaseD_L' + str(i)), size=(5, 1),
                     key='Load_CaseD_L' + str(i), enable_events=True)]]

        Canvas += [
            [sg.Column(COL1), sg.Column(COL2), sg.Column(COL3), sg.Column(COL4), sg.Column(COL5), sg.Column(COL6),
             sg.Column(COL7), sg.Column(COL8), sg.Column(COL9), sg.Column(COL10), sg.Column(COL11)]]
        Canvas += [[sg.Canvas(key='controls_cv' + str(i))], [sg.Canvas(key='Graph' + str(i))]]
    # except:
    # Load_Cases = 0;

    if KeyCheck(variable, 'Cantilever') == True:
        Layout = [[sg.Column([[sg.Column(Layout, vertical_alignment='t'), sg.VSeparator(),
                               sg.Column(Cantilever_Layout(variable), vertical_alignment='t'), sg.VSeparator(),
                               sg.Column(
                                   Canvas,
                                   vertical_alignment='t')]], scrollable=True, expand_x=True,
                             expand_y=True, key='Column')]]
    else:
        Layout = [[sg.Column([[sg.Column(Layout, vertical_alignment='t'), sg.VSeparator(), sg.Column(
            Canvas, vertical_alignment='t')]], scrollable=True, expand_x=True,
                             expand_y=True, key='Column')]]
    return Layout

def steel_calculator(variable):
    # load in data to be read
    from openpyxl import load_workbook

    wb = load_workbook(filename='Steel Design Calculator.xlsx')
    # SectionType = wb.sheetnames
    SectionSize = []
    # print(sheet_ranges['A6'].value)
    SectionType = ['Universal_Beam', 'Universal_Column', 'PFC', 'RHS', 'SHS', 'CHS', 'T-section']
    for j in wb['Universal_Beam']:
        SectionSize.append(j[0].value)
    SectionSize = list(filter(lambda item: item is not None, SectionSize))
    print(SectionSize)
    for j1 in range(4):
        SectionSize.pop(0)
    print(SectionSize)
    # load in GUI library

    sg.theme('GrayGrayGray')
    # ttk_style = 'vista'
    layout = [
        [sg.Column([
            [sg.Text('Input segment Length in metres below:')],
            [sg.Input(default_text=KeyCheck(variable, 'segment length'), key='segment length', size=(5, 1))],
            [sg.Text('Input alpha m below:', key='t3')],
            [sg.Input(key='alpha_m', size=(5, 1), default_text=KeyCheck(variable, 'alpha_m'))],
            [sg.Combo(['FF', 'FP', 'FL', 'PP', 'PL', 'LL', 'FU', 'PU'], key='restraint', default_value='FF',
                      enable_events=True)],
            [sg.Combo(['Shear centre', 'Top flange'], key='load_height_position', default_value='Shear centre',
                      size=(15, 1))],
            [sg.Combo(['Within segment', 'At segment end'], key='longitudinal_position', default_value='Within segment',
                      size=(15, 1)), sg.Text('Load height')],
            [sg.Combo(['Any', 'None', 'One', 'Both'], key='ends_with_restraint', default_value='One'),
             sg.Text('Ends with Lateral restraint')],
            [sg.Button('Calculate Steel', key='calculate', use_ttk_buttons=True)],
            [sg.Column([
                [sg.Text('\u03A6Msx = '), sg.Text('', key='PhiMsx')],
                [sg.Text('\u03A6Mbx = '), sg.Text('', key='PhiMbx')],
                [sg.Text('\u03A6Msy = '), sg.Text('', key='PhiMsy')],
                [sg.Text('\u03A6Mby = '), sg.Text('', key='PhiMby')]
            ]), sg.Column([
                [sg.Text('\u03A6Nsx = '), sg.Text('', key='PhiNsx')],
                [sg.Text('\u03A6Ncx = '), sg.Text('', key='PhiNcx')],
                [sg.Text('\u03A6Nsy = '), sg.Text(key='PhiNsy')],
                [sg.Text('\u03A6Ncy = '), sg.Text(key='PhiNcy')]

            ]), sg.Column([
                [sg.Text('\u03A6Vu = '), sg.Text('', key='PhiVu')],
                [sg.Text('\u03A6Vvm = '), sg.Text('', key='PhiVvm')]
            ])]])]]

    # Create the window
    return layout