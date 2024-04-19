# This script will calculate design actions for use in simple beam problems

# It will determine the moment and shear of a simply supported beam
# It will determine the required Second moment of inertia to meet deflection criteria
import scipy.integrate as integrate
import scipy.special as special
import numpy as np
import os
import matplotlib.pyplot as plt
import matplotlib
import pickle as pickle
import PySimpleGUI as sg
import csv
import Output
import steel_functions as st
import Timber as Timber1
from pylatex import Document, Section, Subsection, Tabular, Math, TikZ, Axis, \
    Plot, Figure, Matrix, Alignat, MultiRow, MultiColumn
from pylatex.utils import italic, NoEscape, bold
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import layouts

class Load_Case:
    def __init__(self, Name, Length, sample_rate, E, I, values):
        self.Name = Name
        self.Length = Length  # In metres
        self.Moment = [0] * round(Length * sample_rate + 1)
        self.Shear = [0] * round(Length * sample_rate + 1)
        self.Deflection = [0] * round(Length * sample_rate + 1)
        try:
            if values['CLength'] > 0:
                self.Deflection = [0] * round(values['CLength'] * sample_rate + Length * sample_rate + 1)
                self.Moment = [0] * round(values['CLength'] * sample_rate + Length * sample_rate + 1)
                self.Shear = [0] * round(values['CLength'] * sample_rate + Length * sample_rate + 1)
        except:
            pass
        self.R1 = 0
        self.R2 = 0
        self.sample_rate = sample_rate
        self.Marea = [0] * round(Length * sample_rate + 1)
        self.E = E
        self.I = I
        self.values = values

    def __str__(self):
        return f"{self.R1} {self.R2} {self.sample_rate}"

    def UDL(self, a, b, w):
        chec = round(self.Length * self.sample_rate + 1)
        Moment = [0] * round(self.Length * self.sample_rate + 1)
        Shear = [0] * round(self.Length * self.sample_rate + 1)
        Deflection = [0] * round(self.Length * self.sample_rate + 1)
        Marea = [0] * round(self.Length * self.sample_rate + 1)
        c = self.Length - b
        b = b - a
        R1 = w * b / 2 / self.Length * (2 * c + b)
        R2 = w * b / 2 / self.Length * (2 * a + b)
        for i in range(round(self.Length * self.sample_rate + 1)):
            if i <= a * self.sample_rate:
                Shear[i] = R1
                Moment[i] = R1 * i / self.sample_rate
            elif i > a * self.sample_rate and i < (a + b) * self.sample_rate:
                Shear[i] = R1 - w * (i / self.sample_rate - a)
                Moment[i] = R1 * i / self.sample_rate - w / 2 * (i / self.sample_rate - a) ** 2
            elif i >= (a + b) * self.sample_rate:
                Shear[i] = -R2
                Moment[i] = R2 * (self.Length - i / self.sample_rate)
                # print(i)
        for i in range(round(self.Length * self.sample_rate)):
            Marea[i] = np.trapz(np.array(Moment[:i]), dx=1 / self.sample_rate)
        centroid = min(range(len(Marea)), key=lambda x: abs(Marea[x] - max(Marea) / 2))
        x1 = lambda x: R1 * x ** 2 / 2
        y1 = integrate.quad(x1, 0, a)
        x1a = lambda x: R1 * x
        y1a = integrate.quad(x1a, 0, a)
        x2 = lambda x: (R1 * x - w / 2 * (x - a) ** 2) * x
        x2a = lambda x: R1 * x - w / 2 * (x - a) ** 2
        y2 = integrate.quad(x2, a, a + b)
        y2a = integrate.quad(x2a, a, a + b)
        x3 = lambda x: R2 * (self.Length - x) * x
        x3a = lambda x: R2 * (self.Length - x)
        y3 = integrate.quad(x3, a + b, self.Length)
        y3a = integrate.quad(x3a, a + b, self.Length)
        try:
            End_deflection = (self.Length - (y1[0] + y2[0] + y3[0]) / (y1a[0] + y2a[0] + y3a[0])) * (
                        y1a[0] + y2a[0] + y3a[0])
            End_deflection1 = Marea[int(self.Length * self.sample_rate)] * (self.Length - centroid / self.sample_rate)
            print(End_deflection1, End_deflection)
            for i in range(round(self.Length * self.sample_rate)):
                if i == 0:
                    Deflection[i] = 0
                elif i < a * self.sample_rate:
                    y1 = integrate.quad(x1, 0, i / self.sample_rate)
                    y1a = integrate.quad(x1a, 0, i / self.sample_rate)

                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                                i / self.sample_rate - (y1[0]) / (y1a[0])) * (y1a[0])) / (self.E * self.I) * 10 ** 6

                elif i < a * self.sample_rate + b * self.sample_rate and i >= a * self.sample_rate:
                    y1 = integrate.quad(x1, 0, a)
                    y1a = integrate.quad(x1a, 0, a)
                    y2 = integrate.quad(x2, a, i / self.sample_rate)
                    y2a = integrate.quad(x2a, a, i / self.sample_rate)
                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                                i / self.sample_rate - (y1[0] + y2[0]) / (y1a[0] + y2a[0])) * (y1a[0] + y2a[0])) / (
                                                self.E * self.I) * 10 ** 6
                elif i >= a * self.sample_rate + b * self.sample_rate:
                    y1 = integrate.quad(x1, 0, a)
                    y1a = integrate.quad(x1a, 0, a)
                    y2 = integrate.quad(x2, a, a + b)
                    y2a = integrate.quad(x2a, a, a + b)
                    y3 = integrate.quad(x3, a + b, i / self.sample_rate)
                    y3a = integrate.quad(x3a, a + b, i / self.sample_rate)
                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                                i / self.sample_rate - (y1[0] + y2[0] + y3[0]) / (y1a[0] + y2a[0] + y3a[0])) * (
                                                 y1a[0] + y2a[0] + y3a[0])) / (self.E * self.I) * 10 ** 6
        except:
            pass
        for i in range(round(self.Length * self.sample_rate + 1)):
            self.Moment[i] += -Moment[i]
            self.Shear[i] += Shear[i]
            self.Deflection[i] += -Deflection[i]
            self.Marea[i] += Marea[i]
        try:

            for i in range(round(self.Length * self.sample_rate + 1),
                           round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):
                self.Deflection[i] += -(i / (self.Length * self.sample_rate) * End_deflection - (
                            i / self.sample_rate - (y1[0] + y2[0] + y3[0]) / (y1a[0] + y2a[0] + y3a[0])) * (
                                                    y1a[0] + y2a[0] + y3a[0])) / (self.E * self.I) * 10 ** 6
        except:
            pass
        self.R1 += R1
        self.R2 += R2

    def Cantilever_UDL(self, a, b, w):
        b = b - a
        Moment = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        Shear = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        Deflection = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        R1 = w * b * (a + b / 2) / self.Length
        R2 = R1 + w * b
        x1a = lambda x: R1 * x
        x1 = lambda x: R1 * x ** 2
        y1 = integrate.quad(x1, 0, self.Length)
        y1a = integrate.quad(x1a, 0, self.Length)
        x2a = lambda x: R1 * x - R2 * (x - self.Length)
        x2 = lambda x: (R1 * x - R2 * (x - self.Length)) * x
        y2 = integrate.quad(x2, self.Length, self.Length + a)
        y2a = integrate.quad(x2a, self.Length, self.Length + a)
        x3a = lambda x: R1 * x - R2 * (x - self.Length) + w * (x - self.Length - a) ** 2 / 2
        x3 = lambda x: (R1 * x - R2 * (x - self.Length) + w * (x - self.Length - a) ** 2 / 2) * x
        y3 = integrate.quad(x3, self.Length + a, self.Length + a + b)
        y3a = integrate.quad(x3a, self.Length + a, self.Length + a + b)
        try:
            End_deflection = (self.Length - (y1[0]) / (y1a[0])) * (
                y1a[0])
            for i in range(round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):
                if i == 0:
                    Moment[i] = R1 * (i / self.sample_rate)
                    Shear[i] = R1
                elif i <= self.Length * self.sample_rate:
                    Moment[i] = R1 * (i / self.sample_rate)
                    Shear[i] = R1
                    y1 = integrate.quad(x1, 0, i / self.sample_rate)
                    y1a = integrate.quad(x1a, 0, i / self.sample_rate)
                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                                i / self.sample_rate - (y1[0]) / (y1a[0])) * (y1a[0])) / (self.E * self.I) * 10 ** 6
                elif i > self.Length * self.sample_rate and i <= self.Length * self.sample_rate + a * self.sample_rate:
                    Moment[i] = R1 * (i / self.sample_rate) - R2 * (i / self.sample_rate - self.Length)
                    Shear[i] = R1 - R2
                    y1 = integrate.quad(x1, 0, self.Length)
                    y1a = integrate.quad(x1a, 0, self.Length)
                    y2 = integrate.quad(x2, self.Length, i / self.sample_rate)
                    y2a = integrate.quad(x2a, self.Length, i / self.sample_rate)
                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                            i / self.sample_rate - (y1[0] + y2[0]) / (y1a[0] + y2a[0])) * (y1a[0] + y2a[0])) / (
                                            self.E * self.I) * 10 ** 6
                elif i > self.Length * self.sample_rate + a * self.sample_rate and i <= self.Length * self.sample_rate + a * self.sample_rate + b * self.sample_rate:
                    Moment[i] = R1 * (i / self.sample_rate) - R2 * (i / self.sample_rate - self.Length) + w * (
                                i / self.sample_rate - self.Length - a) ** 2 / 2
                    Shear[i] = R1 - R2 + w * (i / self.sample_rate - self.Length - a)
                    y1 = integrate.quad(x1, 0, self.Length)
                    y1a = integrate.quad(x1a, 0, self.Length)
                    y2 = integrate.quad(x2, self.Length, a + self.Length)
                    y2a = integrate.quad(x2a, self.Length, a + self.Length)
                    y3 = integrate.quad(x3, a + self.Length, i / self.sample_rate)
                    y3a = integrate.quad(x3a, a + self.Length, i / self.sample_rate)
                    Deflection[i] = (i / (self.Length * self.sample_rate) * End_deflection - (
                            i / self.sample_rate - (y1[0] + y2[0] + y3[0]) / (y1a[0] + y2a[0] + y3a[0])) * (
                                             y1a[0] + y2a[0] + y3a[0])) / (self.E * self.I) * 10 ** 6
        except:
            pass
        for i in range(round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):
            self.Moment[i] += Moment[i]
            self.Shear[i] += -Shear[i]
            self.Deflection[i] += Deflection[i]
        self.R1 += -R1
        self.R2 += R2

    def Point_load(self, a, P):
        b = self.Length - a
        Moment = [0] * round(self.Length * self.sample_rate + 1)
        Shear = [0] * round(self.Length * self.sample_rate + 1)
        Deflection = [0] * round(self.Length * self.sample_rate + 1)
        Marea = [0] * round(self.Length * self.sample_rate + 1)
        R1 = P * b / self.Length
        R2 = P * a / self.Length
        for i in range(round(self.Length * self.sample_rate + 1)):
            if a == 0 or b == 0:
                Moment[i] = P * b * (i / self.sample_rate) / self.Length
                Shear[i] = R1
            elif b == 0:
                Moment[i] = P * a * (self.Length - i / self.sample_rate) / self.Length
                Shear[i] = -R2
            else:
                if i < a * self.sample_rate:
                    Moment[i] = P * b * (i / self.sample_rate) / self.Length
                    Shear[i] = R1
                    Deflection[i] = P * b * (i / self.sample_rate) / (6 * self.Length * self.E * self.I) * (
                                self.Length ** 2 - b ** 2 - (i / self.sample_rate) ** 2) * 10 ** 6
                else:
                    Moment[i] = P * a * (self.Length - i / self.sample_rate) / self.Length
                    Shear[i] = -R2
                    Deflection[i] = P * a * (self.Length - i / self.sample_rate) / (
                                6 * self.Length * self.E * self.I) * (
                                            self.Length * 2 * (i / self.sample_rate) - a ** 2 - (
                                                i / self.sample_rate) ** 2) * 10 ** 6
        for i in range(round(self.Length * self.sample_rate + 1)):
            self.Moment[i] += -Moment[i]
            self.Shear[i] += Shear[i]
            self.Deflection[i] += -Deflection[i]
            self.Marea[i] += Marea[i]
            try:

                for i in range(round(self.Length * self.sample_rate + 1),
                               round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):

                    self.Deflection[i] += (P*a*(self.Length - a)* (i/self.sample_rate - self.Length)) /(6*self.E * self.I * self.Length)*(self.Length + a)*10**3
            except:
                pass
        self.R1 += R1
        self.R2 += R2

    def Cantilever_Point_load(self, a, P):
        Moment = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        Shear = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        Deflection = [0] * round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)
        R1 = P * a / self.Length
        R2 = P / self.Length * (self.Length + a)
        try:
            for i in range(round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):
                if i <= self.Length * self.sample_rate:
                    Moment[i] = P * a * (i / self.sample_rate) / self.Length
                    Shear[i] = R1
                    Deflection[i] = (P * a * i / self.sample_rate) / (6 * self.Length * self.E * self.I) * (
                                self.Length ** 2 - (i / self.sample_rate) ** 2) * 10 ** 6
                elif i > self.Length * self.sample_rate and i <= self.Length * self.sample_rate + a * self.sample_rate:
                    Moment[i] = P * (a - (i / self.sample_rate - self.Length))
                    Shear[i] = -R2 + R1
                    Deflection[i] = -(P * (i / self.sample_rate - self.Length)) / (6 * self.E * self.I) * (
                                2 * a * self.Length + 3 * a * (i / self.sample_rate - self.Length) - (
                                    i / self.sample_rate - self.Length) ** 2) * 10 ** 6
                    Deflect = Deflection[i]
                elif i > self.Length * self.sample_rate + a * self.sample_rate:
                    Deflection[i] = (i / self.sample_rate - self.Length) / a * Deflect
        except:
            pass
        for i in range(round(self.Length * self.sample_rate + self.values['CLength'] * self.sample_rate + 1)):
            self.Moment[i] += Moment[i]
            self.Shear[i] += -Shear[i]
            self.Deflection[i] += Deflection[i]
        self.R1 += -R1
        self.R2 += R2


class ORMType:
    def __init__(self, key, value):
        self.key = value

    def __repr__(self):
        return "ok"


class Beam:
    def __init__(self, Name, Length, sample_rate, E, Ix, Iy, Loadings, Cpe_V, Cpi_V, Cpe_H, Cpi_H, P, values,
                 Cantilever):
        self.Name = Name
        self.E = E
        self.Ix = Ix
        self.Iy = Iy
        self.Length = Length
        self.sample_rate = sample_rate
        self.Loadings = Loadings
        self.Cpe_V = Cpe_V
        self.Cpi_V = Cpi_V
        self.Cpe_H = Cpe_H
        self.Cpi_H = Cpi_H
        self.P = P
        self.values = values
        self.Cantilever = Cantilever
        self.MaxDeflections = {}
        self.Combined_Deflections ={}
        self.section_properties = {}

    def __str__(self):
        return self.Name

    def LoadCase(self, Loadings, Cantilever):

        for key in Loadings:
            if key == 'WOoP':
                setattr(self, key, Load_Case(self.Name, self.Length, self.sample_rate, self.E, self.Iy, self.values))
            else:
                setattr(self, key, Load_Case(self.Name, self.Length, self.sample_rate, self.E, self.Ix, self.values))
            # self.key = Load_Case(self.name, self.Length, self.sample_rate, self.E, self.I)
            # try:
            for j in range(len(Loadings[key][0])):
                getattr(self, key).UDL(Loadings[key][0][j][0], Loadings[key][0][j][1], Loadings[key][0][j][2])
                # self.key.UDL(Loadings[key][0][j][0],Loadings[key][0][j][1],Loadings[key][0][j][2])
            # except:
            # pass
            # try:
            for j in range(len(Loadings[key][1])):
                getattr(self, key).Point_load(Loadings[key][1][j][0], Loadings[key][1][j][1])
            # except:
            # pass
            if self.values['Cantilever'] == True:

                for j in range(len(Cantilever[key][0])):
                    getattr(self, key).Cantilever_UDL(Cantilever[key][0][j][0], Cantilever[key][0][j][1],
                                                      Cantilever[key][0][j][2])
                    # self.key.UDL(Loadings[key][0][j][0],Loadings[key][0][j][1],Loadings[key][0][j][2])
                # except:
                # pass
                # try:
                for j in range(len(Cantilever[key][1])):
                    getattr(self, key).Cantilever_Point_load(Cantilever[key][1][j][0], Cantilever[key][1][j][1])

def open_file(Name, Project):
    try:
        with open(Project + '\\' + Name + '.pkl', 'rb') as inp:
            foo = pickle.load(inp)
            return foo
    except (OSError, IOError) as e:
        print('No info')


def Method(Name, Length, sample_rate, E, I, Iy, Loadings, Cpe_V, Cpi_V, Cpe_H, Cpi_H, P, values, Cantilever):
    Name1 = Beam(Name, Length, sample_rate, E, I, Iy, Loadings, Cpe_V, Cpi_V, Cpe_H, Cpi_H, P, values, Cantilever)
    Name1.LoadCase(Loadings, Cantilever)

    with open(values['Project'] + '\\' + Name + '.pkl', 'wb') as outp:
        pickle.dump(Name1, outp, pickle.HIGHEST_PROTOCOL)


def variable_write(values, Name):
    # if values['Member Type'] == 'Beam':
    with open(Name + '.txt', 'w', newline='') as csv_file:
        data = [[str(i), values[i]] for i in values]
        print(data)
        writer = csv.writer(csv_file)
        writer.writerows(data)


def draw_figure_w_toolbar(canvas, fig, canvas_toolbar):
    if canvas.children:
        for child in canvas.winfo_children():
            child.destroy()
    if canvas_toolbar.children:
        for child in canvas_toolbar.winfo_children():
            child.destroy()
    figure_canvas_agg = FigureCanvasTkAgg(fig, master=canvas)
    figure_canvas_agg.draw()
    toolbar = Toolbar(figure_canvas_agg, canvas_toolbar)
    toolbar.update()
    figure_canvas_agg.get_tk_widget().pack(side='right', fill='both', expand=1)


class Toolbar(NavigationToolbar2Tk):
    def __init__(self, *args, **kwargs):
        super(Toolbar, self).__init__(*args, **kwargs)


def variables(Name):
    Dictionary = {}
    with open(str(Name) + '.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            Dictionary[row[0]] = row[1]
    return Dictionary






def delete_figure_agg(figure_agg):
    figure_agg.get_tk_widget().forget()
    plt.close('all')


def draw_figure(canvas, figure):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    return figure_canvas_agg


def draw_graph(Beam,i):
    # fig = plt.figure(figsize=(12,30),dpi=80)
    fig = matplotlib.figure.Figure(figsize=(12, 4),dpi = 80)
    try:
        # Length = np.arange(0,Beam.Length + Beam.values['CLength']+ 1/Beam.sample_rate,1/Beam.sample_rate)
        Length = range(round(Beam.Length * Beam.sample_rate + Beam.values['CLength'] * Beam.sample_rate + 1))
        Length = [i / Beam.sample_rate for i in Length]
        Zeros = [0] * (int(Beam.Length + Beam.values['CLength']) * Beam.sample_rate + 1)
    except:

        # Length =  np.arange(0,Beam.Length + 1/Beam.sample_rate,1/Beam.sample_rate)
        Length = range(round(Beam.Length * Beam.sample_rate + 1))
        Length = [i / Beam.sample_rate for i in Length]



    j = 0
    R1 = float(Beam.values['Load_CaseG1'+str(i)])* getattr(Beam,'G1').R1 + float(Beam.values['Load_CaseG2'+str(i)])* getattr(Beam,'G2').R1 + float(Beam.values['Load_CaseQ'+str(i)])* getattr(Beam,'Q').R1 - float(Beam.values['Load_CaseWup'+str(i)])* getattr(Beam,'Wup').R1 + float(Beam.values['Load_CaseWdown'+str(i)])* getattr(Beam,'Wdown').R1 + float(Beam.values['Load_CaseWOoP'+str(i)])* getattr(Beam,'WOoP').R1
    R2 = float(Beam.values['Load_CaseG1'+str(i)])* getattr(Beam,'G1').R2 + float(Beam.values['Load_CaseG2'+str(i)])* getattr(Beam,'G2').R2 + float(Beam.values['Load_CaseQ'+str(i)])* getattr(Beam,'Q').R2 - float(Beam.values['Load_CaseWup'+str(i)])* getattr(Beam,'Wup').R2 + float(Beam.values['Load_CaseWdown'+str(i)])* getattr(Beam,'Wdown').R2 + float(Beam.values['Load_CaseWOoP'+str(i)])* getattr(Beam,'WOoP').R2
    if Beam.values['Load_CaseM'+str(i)] ==True:
        j += 1
        Combined_Moment = [float(Beam.values['Load_CaseG1'+str(i)])* x1 + float(Beam.values['Load_CaseG2'+str(i)])* x2 + float(Beam.values['Load_CaseQ'+str(i)])* x3 - float(Beam.values['Load_CaseWup'+str(i)])* x4 + float(Beam.values['Load_CaseWdown'+str(i)])* x5 + float(Beam.values['Load_CaseWOoP'+str(i)])* x6 for x1, x2, x3, x4, x5, x6 in zip(getattr(Beam, 'G1').Moment,getattr(Beam, 'G2').Moment,getattr(Beam, 'Q').Moment,getattr(Beam, 'Wup').Moment,getattr(Beam, 'Wdown').Moment,getattr(Beam, 'WOoP').Moment)]


    if Beam.values['Load_CaseS' + str(i)] == True:
        j +=1
        Combined_Shear = [float(Beam.values['Load_CaseG1'+str(i)]) * x1 + float(Beam.values['Load_CaseG2'+str(i)]) * x2 + float(
            Beam.values['Load_CaseQ'+str(i)]) * x3 - float(Beam.values['Load_CaseWup'+str(i)]) * x4 + float(
            Beam.values['Load_CaseWdown'+str(i)]) * x5 + float(Beam.values['Load_CaseWOoP'+str(i)]) * x6 for x1, x2, x3, x4, x5, x6 in
                    zip(getattr(Beam, 'G1').Shear, getattr(Beam, 'G2').Shear,getattr(Beam, 'Q').Shear,
                        getattr(Beam, 'Wup').Shear, getattr(Beam, 'Wdown').Shear, getattr(Beam, 'WOoP').Shear)]
    if Beam.values['Load_CaseD' + str(i)] == True:
        j += 1
        Combined_Deflection = [float(Beam.values['Load_CaseG1'+str(i)]) * x1 + float(Beam.values['Load_CaseG2'+str(i)]) * x2 + float(
            Beam.values['Load_CaseQ'+str(i)]) * x3 - float(Beam.values['Load_CaseWup'+str(i)]) * x4 + float(
            Beam.values['Load_CaseWdown'+str(i)]) * x5 + float(Beam.values['Load_CaseWOoP'+str(i)]) * x6 for x1, x2, x3, x4, x5, x6 in
                    zip(getattr(Beam, 'G1').Deflection, getattr(Beam, 'G2').Deflection,getattr(Beam, 'Q').Deflection,
                        getattr(Beam, 'Wup').Deflection, getattr(Beam, 'Wdown').Deflection, getattr(Beam, 'WOoP').Deflection)]

        # Factor deflections for timber creep factors
        if Beam.values['SectionType'] == 'MGP 10' or Beam.values['SectionType'] == 'MGP 12':
            if Beam.values['j2'] == '<15%':
                Combined_Deflection = [2 * i for i in Combined_Deflection]

            else:
                Combined_Deflection = [3 * i for i in Combined_Deflection]
        # Find the Maximum deflection from the load case selected and add to the Beam object

        getattr(Beam, 'MaxDeflections')[i] = max(Combined_Deflection, key=abs)
    j1 = 1
    if Beam.values['Load_CaseM'+str(i)] ==True:
        ax1 = fig.add_subplot(1, j, j1)
        ax1.set_ylabel('Moment [KNm]')
        ax1.set_xlabel('Length [m]')
        ax1.set_title(Beam.values['Load_Case'+str(i)])
        ax1.plot(Length, Combined_Moment)
        ax1.text(0, -abs(1 / 10 * max(Combined_Moment, key=abs)), str(round(R1, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.text(Beam.Length, -abs(1 / 10 * max(Combined_Moment, key=abs)), str(round(R2, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.plot([0, Beam.Length], [0, 0], 'g', marker=6, markersize=15)
        try:
            ax1.plot([0, Beam.Length], [-Beam.section_properties[0][0]/1000, -Beam.section_properties[0][0]/1000], '--r')
        except:
            pass
        try:
            ax1.plot([0, Beam.Length], [-Beam.section_properties[0]['PhiMbx'], -Beam.section_properties[0]['PhiMbx']], '--r')
        except:
            pass
        try:
            ax1.plot([Beam.Length, Beam.Length + Beam.values['CLength']], [0, 0], 'g')
        except:
            pass
        j1 += 1

    if Beam.values['Load_CaseS' + str(i)] == True:
        ax1 = fig.add_subplot(1, j, j1)
        ax1.set_ylabel('Shear [KN]')
        ax1.set_xlabel('Length [m]')
        ax1.set_title(Beam.values['Load_Case'+str(i)])
        ax1.plot(Length, Combined_Shear)
        ax1.text(0, -abs(1 / 10 * max(Combined_Shear, key=abs)), str(round(R1, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.text(Beam.Length, -abs(1 / 10 * max(Combined_Shear, key=abs)), str(round(R2, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.plot([0, Beam.Length], [0, 0], 'g', marker=6, markersize=15)
        try:
            ax1.plot([0, Beam.Length], [-Beam.section_properties[1][0]/1000, -Beam.section_properties[1][0]/1000], '--r')
        except:
            pass
        try:
            ax1.plot([0, Beam.Length], [-Beam.section_properties[0]['PhiVu'], -Beam.section_properties[0]['PhiVu']], '--r')
        except:
            pass
        try:
            ax1.plot([Beam.Length, Beam.Length + Beam.values['CLength']], [0, 0], 'g')
        except:
            pass
        j1 +=1

    if Beam.values['Load_CaseD' + str(i)] == True:
        ax1 = fig.add_subplot(1, j, j1)
        ax1.set_ylabel('Deflection [mm]')
        ax1.set_xlabel('Length [m]')
        ax1.set_title(Beam.values['Load_Case'+str(i)])
        ax1.plot(Length, Combined_Deflection)
        ax1.text(0, -abs(1 / 10 * max(Combined_Deflection, key=abs)), str(round(R1, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.text(Beam.Length, -abs(1 / 10 * max(Combined_Deflection, key=abs)), str(round(R2, 1)) + 'KN',
                 verticalalignment='top', horizontalalignment='center')
        ax1.plot([0, Beam.Length], [0, 0], 'g', marker=6, markersize=15)
        ax1.plot([0, Beam.Length], [-max(Length)/Beam.values['Load_CaseD_L'+str(i)]*1000, -max(Length)/Beam.values['Load_CaseD_L'+str(i)]*1000], '--r')
        try:
            ax1.plot([Beam.Length, Beam.Length + Beam.values['CLength']], [0, 0], 'g')
            ax1.plot([Beam.Length, Beam.Length + Beam.values['CLength']], [-Beam.values['CLength'] / 0.180, -Beam.values['CLength'] / 0.180], '--r')
        except:
            pass
        j1 += 1

    fig.tight_layout()
    getattr(Beam,'Combined_Deflections')[i] = fig
    return fig

def PassFail(Beam):
    String = ''
    try:
        # Length = np.arange(0,Beam.Length + Beam.values['CLength']+ 1/Beam.sample_rate,1/Beam.sample_rate)
        Length = range(round(Beam.Length * Beam.sample_rate + Beam.values['CLength'] * Beam.sample_rate + 1))
        Length = [i / Beam.sample_rate for i in Length]
        Zeros = [0] * (int(Beam.Length + Beam.values['CLength']) * Beam.sample_rate + 1)
    except:

        # Length =  np.arange(0,Beam.Length + 1/Beam.sample_rate,1/Beam.sample_rate)
        Length = range(round(Beam.Length * Beam.sample_rate + 1))
        Length = [i / Beam.sample_rate for i in Length]
    for i in range(int(Beam.values['Load_Cases'])):
        try:
            if abs(max(Length) / Beam.values['Load_CaseD_L' + str(i)] * 1000) >= abs(Beam.MaxDeflections[i]):
                String += Beam.values['Load_Case'+str(i)] + ' = Pass | Limit = ' + str(round(abs(max(Length) / Beam.values['Load_CaseD_L' + str(i)] * 1000),1)) +'mm > '+ str(round(abs(Beam.MaxDeflections[i]),1)) +'mm\n'
            else:
                String += Beam.values['Load_Case'+str(i)] + ' = Fail | Limit = '+ str(round(abs(max(Length) / Beam.values['Load_CaseD_L' + str(i)] * 1000),1)) +'mm < '+ str(round(abs(Beam.MaxDeflections[i]),1)) + 'mm\n'
        except:
            continue
    return String
def Layouts(variable):
    Layout = layouts.Layouts(variable)
    window = sg.Window('Window', Layout, resizable=True).finalize()
    window.Maximize()
    from openpyxl import load_workbook
    sg.theme('GrayGrayGray')
    wb = load_workbook(filename='Steel Design Calculator.xlsx')
    # SectionType = wb.sheetnames
    SectionSize = []
    # print(sheet_ranges['A6'].value)
    SectionType = ['Universal_Beam', 'Universal_Column', 'PFC', 'RHS', 'SHS', 'CHS', 'T-section']
    for j in wb['Universal_Beam']:
        SectionSize.append(j[0].value)
    SectionSize = list(filter(lambda item: item is not None, SectionSize))
    print(SectionSize)
    for j1 in range(4):
        SectionSize.pop(0)
    fig_canvas_agg = None
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Quit':
            break
        for i in values:
            if event == i:
                variable_write(values, 'Loading')
        if event == 'UDLS' or event == 'Point_loads' or event == 'Ex_Point_loads' or event == 'CUDLS' or event == 'CPoint_loads' or event == 'CEx_Point_loads' or event == 'Cantilever':
            window.close()
            variable = variables('Loading')
            Layouts(variable)
        if event == 'Default_Load_Cases':

            variable_write(values, 'Loading')
            window.close()

            variable = variables('Loading')
            for i in variable:
                if 'Load_Case' in i:
                    variable[i] = '0'
            variable['Load_Cases'] ='10'
            variable['Load_Case0'] = '1.2G2 + 1.5Q'
            variable['Load_CaseG20']='1.2'
            variable['Load_CaseQ0']='1.5'
            variable['Load_CaseM0'] = '1'
            variable['Load_CaseS0'] = '1'

            variable['Load_Case1'] = '1.35G2'
            variable['Load_CaseG21']='1.35'
            variable['Load_CaseM1'] = '1'
            variable['Load_CaseS1'] = '1'

            variable['Load_Case2'] = '0.9G1 + 1.5Wup'
            variable['Load_CaseG12']='0.9'
            variable['Load_CaseWup2']='1.5'
            variable['Load_CaseM2'] = '1'
            variable['Load_CaseS2'] = '1'

            variable['Load_Case3'] = '1.2G2 + 1.5Wdown'
            variable['Load_CaseG23'] = '1.2'
            variable['Load_CaseWdown3'] = '1.5'
            variable['Load_CaseM3'] = '1'
            variable['Load_CaseS3'] = '1'

            variable['Load_Case4'] = '1.5WOoP'
            variable['Load_CaseWOoP4']='1.5'
            variable['Load_CaseM4'] = '1'
            variable['Load_CaseS4'] = '1'

            variable['Load_Case5'] = 'G2'
            variable['Load_CaseG25']='1'
            variable['Load_CaseD5'] = '1'
            variable['Load_CaseD_L5'] = '360'

            variable['Load_Case6'] = 'Q'
            variable['Load_CaseQ6']='1'
            variable['Load_CaseD6'] = '1'
            variable['Load_CaseD_L6'] = '250'

            variable['Load_Case8'] = 'Wdown'
            variable['Load_CaseWdown8']='1'
            variable['Load_CaseD8'] = '1'
            variable['Load_CaseD_L8'] = '150'

            variable['Load_Case7'] = 'Wup'
            variable['Load_CaseWup7']='1'
            variable['Load_CaseD7'] = '1'
            variable['Load_CaseD_L7'] = '150'

            variable['Load_Case9'] = 'WOoP'
            variable['Load_CaseWOoP9']='1'
            variable['Load_CaseD9'] = '1'
            variable['Load_CaseD_L9'] = '150'
            Layouts(variable)
        if event == 'Length_Update':
            variable_write(values, 'Loading')
            for i in range(int(variable['UDLS'])):
                window['segment length'].update(values['Length'])
                window['G1b' + str(i)].update(values['Length'])
                window['G2b' + str(i)].update(values['Length'])
                window['Qb' + str(i)].update(values['Length'])
                window['Wupb' + str(i)].update(values['Length'])
                window['Wdownb' + str(i)].update(values['Length'])
                window['WOoPb' + str(i)].update(values['Length'])
        if event == 'CLength_Update':
            variable_write(values, 'Loading')
            for i in range(int(variable['CUDLS'])):
                i = str(i) + 'C'
                window['G1b' + str(i)].update(values['Length'])
                window['G2b' + str(i)].update(values['Length'])
                window['Qb' + str(i)].update(values['Length'])
                window['Wupb' + str(i)].update(values['Length'])
                window['Wdownb' + str(i)].update(values['Length'])
                window['WOoPb' + str(i)].update(values['Length'])
        if event == 'calculate' or event == 'Calculate':
           try:
                SectionType1 = values['SectionType']
                SectionSize1 = values['SectionSize']

                alpha_m = float(values['alpha_m'])
                restraint = values['restraint']
                load_height_position = values['load_height_position']
                longitudinal_position = values['longitudinal_position']
                ends_with_restraint = values['ends_with_restraint']
                section_properties = st.section_properties(SectionType1, SectionSize1, 0, 0, 0)
                st.compact(section_properties, SectionType1, float(values['segment length']), alpha_m, restraint,
                           load_height_position,
                           longitudinal_position, ends_with_restraint)
                st.shear(section_properties, SectionType1)
                st.shear_moment(section_properties, 0)
                st.axial_compression(section_properties, SectionType1, float(values['segment length']))

                print(section_properties)
                try:
                    del Timber12
                    del Timber13
                except:
                    pass
                try:
                    del section_properties1
                except:
                    pass
                section_properties1 = section_properties
                # End program if user closes window or
                # presses the OK button
                # if event == 'quit' or event == sg.WIN_CLOSED:

                window['PhiMsx'].update(str(round(section_properties['PhiMsx'], 1)) + 'kNm')

                window['PhiMbx'].update(str(round(section_properties['PhiMbx'], 1)) + 'kNm')

                window['PhiMsy'].update(str(round(section_properties['PhiMsy'], 1)) + 'kNm')
                window['PhiMby'].update(str(round(section_properties['PhiMby'], 1)) + 'kNm')
                window['PhiNsx'].update(str(round(section_properties['PhiNsx'], 1)) + 'kN')
                window['PhiNcx'].update(str(round(section_properties['PhiNcx'], 1)) + 'kN')
                window['PhiNsy'].update(str(round(section_properties['PhiNsy'], 1)) + 'kN')
                window['PhiNcy'].update(str(round(section_properties['PhiNcy'], 1)) + 'kN')
                window['PhiVu'].update(str(round(section_properties['PhiVu'], 2)) + ' KN')
           except:
            MGP10 = {'90x90': [90, 90], '140x90': [140, 90], '190x90': [190, 90], '240x90': [240, 90],
                        '300x90': [300, 90], '90x45': [90, 45], '140x45': [140, 45], '190x45': [190, 45],
                        '240x45': [240, 45]}
            Timber = MGP10[values['SectionSize']]
            if values['Seasoned'] == 'Seasoned':
                Seasoned = True
            else:
                Seasoned = False
            try:
                del section_properties1
            except:
                pass
            try:
                del Timber12
                del Timber13
            except:
                pass
            Timber12 = Timber1.beam_moment(Timber[1],Timber[0],values['load_duration'],Seasoned,1,1,1,float(values['segment length'])*1000,values['SectionType'],True)
            Timber13 = Timber1.beam_shear(Timber[1],Timber[0],values['load_duration'],Seasoned,1,1,1,float(values['segment length'])*1000,values['SectionType'])
        if event == 'Calculate':
            variable_write(values, 'Loading')
            if fig_canvas_agg:
                delete_figure_agg(fig_canvas_agg)
            plt.close('all')
            G1U = []
            G2U = []
            QU = []
            WupU = []
            WdownU = []
            WOoPU = []
            G1P = []
            G2P = []
            QP = []
            WupP = []
            WdownP = []
            WOoPP = []
            for i in values:
                try:
                    values[i] = float(values[i])
                except:
                    print(values[i])
            for i in range(UDLS):
                G1U += [
                    [values['G1a' + str(i)], values['G1b' + str(i)], values['G1M' + str(i)] * values['G1T' + str(i)]]]
                G2U += [
                    [values['G2a' + str(i)], values['G2b' + str(i)], values['G2M' + str(i)] * values['G2T' + str(i)]]]
                QU += [[values['Qa' + str(i)], values['Qb' + str(i)], values['QM' + str(i)] * values['QT' + str(i)]]]

                WupU += [[values['Wupa' + str(i)], values['Wupb' + str(i)], values['WupM' + str(i)] * values['WupT' + str(i)]]]

                WdownU += [[values['Wdowna' + str(i)], values['Wdownb' + str(i)], values['WdownM' + str(i)] * values['WdownT' + str(i)]]]

                WOoPU += [[values['WOoPa' + str(i)], values['WOoPb' + str(i)], values['WOoPM' + str(i)] * values['WOoPT' + str(i)]]]


            for i in range(Point_loads):
                G1P += [[values['G1Pa' + str(i)], values['G1P' + str(i)]]]
                G2P += [[values['G2Pa' + str(i)], values['G2P' + str(i)]]]
                QP += [[values['QPa' + str(i)], values['QP' + str(i)]]]
                WupP += [[values['WupPa' + str(i)], values['WupP' + str(i)]]]
                WdownP += [[values['WdownPa' + str(i)], values['WdownP' + str(i)]]]
                WOoPP += [[values['WOoPPa' + str(i)], values['WOoPP' + str(i)]]]
            for i in range(Ex_Point_loads):
                Beam = open_file(values['Ex' + str(i)], values['Project'])
                if values['L/R' + str(i)] == 'Left':
                    G1P += [[values['Exa' + str(i)], Beam.G1.R1]]
                    G2P += [[values['Exa' + str(i)], Beam.G2.R1]]
                    QP += [[values['Exa' + str(i)], Beam.Q.R1]]
                    WupP += [[values['Exa' + str(i)], Beam.Wup.R1]]
                    WdownP += [[values['Exa' + str(i)], Beam.Wdown.R1]]
                    WOoPP += [[values['Exa' + str(i)], Beam.WOoP.R1]]
                else:
                    G1P += [[values['Exa' + str(i)], Beam.G1.R2]]
                    G2P += [[values['Exa' + str(i)], Beam.G2.R2]]
                    QP += [[values['Exa' + str(i)], Beam.Q.R2]]
                    WupP += [[values['Exa' + str(i)], Beam.Wup.R2]]
                    WdownP += [[values['Exa' + str(i)], Beam.Wdown.R2]]
                    WOoPP += [[values['Exa' + str(i)], Beam.WOoP.R2]]
            Loadings = {'G1': [G1U, G1P], 'G2': [G2U, G2P], 'Q': [QU, QP], 'Wup': [WupU, WupP],
                        'Wdown': [WdownU, WdownP], 'WOoP': [WOoPU, WOoPP]}
            if values['Ix_check'] == False:
                try:
                    SectionType1 = values['SectionType']
                    SectionSize1 = values['SectionSize']
                    section_properties = st.section_properties(SectionType1, SectionSize1, 0, 0, 0)
                    Ix = section_properties['Ix'] * 10 ** -12
                except:
                    MGP10 = {'90x90': [90, 90], '140x90': [140, 90], '190x90': [190, 90], '240x90': [240, 90],
                             '300x90': [300, 90], '90x45': [90, 45], '140x45': [140, 45], '190x45': [190, 45],
                             '240x45': [240, 45]}
                    Timber = MGP10[values['SectionSize']]
                    Ix = Timber[1] * Timber[0] ** 3 / 12 * 10 ** -12
            else:
                Ix = values['Ix'] * 10 ** -12
            if values['Iy_check'] == False:
                try:
                    SectionType1 = values['SectionType']
                    SectionSize1 = values['SectionSize']
                    section_properties = st.section_properties(SectionType1, SectionSize1, 0, 0, 0)
                    Iy = section_properties['Iy'] * 10 ** -12
                except:
                    MGP10 = {'90x90': [90, 90], '140x90': [140, 90], '190x90': [190, 90], '240x90': [240, 90],
                             '300x90': [300, 90],'90x45':[90,45],'140x45':[140,45],'190x45':[190,45],'240x45':[240,45]}
                    Timber = MGP10[values['SectionSize']]
                    Iy = Timber[0] * Timber[1] ** 3 / 12 * 10 ** -12
            else:
                Iy = values['Iy'] * 10 ** -12

            CG1U = []
            CG2U = []
            CQU = []
            CWupU = []
            CWdownU = []
            CWOoPU = []
            CG1P = []
            CG2P = []
            CQP = []
            CWupP = []
            CWdownP = []
            CWOoPP = []
            try:
                CUDLS = int(float(KeyCheck(variable, 'CUDLS')))
            except:
                CUDLS = 1
            try:
                CPoint_loads = int(float(KeyCheck(variable, 'CPoint_loads')))
            except:
                CPoint_loads = 1
            try:
                CEx_Point_loads = int(float(KeyCheck(variable, 'CEx_Point_loads')))
            except:
                CEx_Point_loads = 1
            for i in values:
                try:
                    values[i] = float(values[i])
                except:
                    print(values[i])
            try:
                for i in range(CUDLS):
                    i = str(i) + 'C'
                    CG1U += [[values['G1a' + str(i)], values['G1b' + str(i)],
                              values['G1M' + str(i)] * values['G1T' + str(i)]]]
                    CG2U += [[values['G2a' + str(i)], values['G2b' + str(i)],
                              values['G2M' + str(i)] * values['G2T' + str(i)]]]
                    CQU += [
                        [values['Qa' + str(i)], values['Qb' + str(i)], values['QM' + str(i)] * values['QT' + str(i)]]]
                    CWupM = (values['Cpe_V'] + values['Cpi_V']) * values['P']
                    CWupU += [[values['Wupa' + str(i)], values['Wupb' + str(i)], CWupM * values['WupT' + str(i)]]]
                    CWdownM = (values['Cpe_V'] + values['Cpi_V']) * values['P']
                    CWdownU += [
                        [values['Wdowna' + str(i)], values['Wdownb' + str(i)], CWdownM * values['WdownT' + str(i)]]]
                    CWOoPM = (values['Cpe_H'] + values['Cpi_H']) * values['P']
                    CWOoPU += [[values['WOoPa' + str(i)], values['WOoPb' + str(i)], CWOoPM * values['WOoPT' + str(i)]]]
                for i in range(CPoint_loads):
                    i = str(i) + 'C'
                    CG1P += [[values['G1Pa' + str(i)], values['G1P' + str(i)]]]
                    CG2P += [[values['G2Pa' + str(i)], values['G2P' + str(i)]]]
                    CQP += [[values['QPa' + str(i)], values['QP' + str(i)]]]
                    CWupP += [[values['WupPa' + str(i)], values['WupP' + str(i)]]]
                    CWdownP += [[values['WdownPa' + str(i)], values['WdownP' + str(i)]]]
                    CWOoPP += [[values['WOoPPa' + str(i)], values['WOoPP' + str(i)]]]
                for i in range(CEx_Point_loads):
                    i = str(i) + 'C'
                    Beam = open_file(values['Ex' + str(i)], values['Project'])
                    if values['L/R' + str(i)] == 'Left':
                        CG1P += [[values['Exa' + str(i)], Beam.G1.R1]]
                        CG2P += [[values['Exa' + str(i)], Beam.G2.R1]]
                        CQP += [[values['Exa' + str(i)], Beam.Q.R1]]
                        CWupP += [[values['Exa' + str(i)], Beam.Wup.R1]]
                        CWdownP += [[values['Exa' + str(i)], Beam.Wdown.R1]]
                        CWOoPP += [[values['Exa' + str(i)], Beam.WOoP.R1]]
                    else:
                        CG1P += [[values['Exa' + str(i)], Beam.G1.R2]]
                        CG2P += [[values['Exa' + str(i)], Beam.G2.R2]]
                        CQP += [[values['Exa' + str(i)], Beam.Q.R2]]
                        CWupP += [[values['Exa' + str(i)], Beam.Wup.R2]]
                        CWdownP += [[values['Exa' + str(i)], Beam.Wdown.R2]]
                        CWOoPP += [[values['Exa' + str(i)], Beam.WOoP.R2]]
                Cantilever = {'G1': [CG1U, CG1P], 'G2': [CG2U, CG2P], 'Q': [CQU, CQP], 'Wup': [CWupU, CWupP],
                              'Wdown': [CWdownU, CWdownP], 'WOoP': [CWOoPU, CWOoPP]}
            except:
                Cantilever = 0
            if values['E_check'] == True:
                E = values['E'] * 10 ** 9
            else:
                if values['SectionType'] == 'MGP 10':
                    E = 10 * 10 ** 9
                elif values['SectionType'] == 'MGP 12':
                    E = 12.7 * 10 ** 9
                else:
                    E = 200 * 10 ** 9
            Method(values['Name'], values['Length'], 100, E, Ix, Iy, Loadings,
                   values['Cpe_V'], values['Cpi_V'], values['Cpe_H'], values['Cpi_H'], values['P'], values, Cantilever)

            Beam1 = open_file(values['Name'], values['Project'])
            try:
                getattr(Beam1,'section_properties')[0] = section_properties1
            except:
                pass
            try:
                getattr(Beam1,'section_properties')[0] = Timber12
                getattr(Beam1, 'section_properties')[1] = Timber13
            except:
                pass
            # fig_canvas_agg = draw_figure(window['ULSGraph'].TKCanvas,draw_graph(Beam1))
            for i in range(Load_Cases):

                draw_figure_w_toolbar(window['Graph'+str(i)].TKCanvas, draw_graph(Beam1,i), window['controls_cv'+str(i)].TKCanvas)

            window['P/F_Check'].update(PassFail(Beam1))
            window.refresh()
            window['Column'].contents_changed()
            Output.Main(Beam1)
        if event == 'Reload':
            try:
                directory = KeyCheck(variable, 'Project')
                existing = []
                for file in os.listdir(directory):
                    filename = os.fsdecode(file)
                    if filename.endswith(".pkl"):
                        filename = os.path.basename(filename)
                        filename = os.path.splitext(filename)[0]
                        existing += [filename]
                        continue
                    else:
                        continue
            except:
                directory = os.fsencode(os.getcwd())
            Layout1 = [
                [sg.Text('Select Existing Beam')],
                [sg.Combo(existing, key='Reload_Ex', enable_events=True)],
                [sg.Button('Continue', key='Reload_continue')]
            ]
            Project = values['Project']
            window1 = sg.Window('Select Existing Beam', Layout1, resizable=True).finalize()
            window.close()
            while True:
                event, values = window1.read()
                if event == sg.WINDOW_CLOSED or event == 'Quit':
                    break
                elif event == 'Reload_continue':
                    Beam = open_file(values['Reload_Ex'], KeyCheck(variable, 'Project'))
                    window1.close()
                    Beam.values['Project'] = Project
                    Layouts(Beam.values)
                    break

        if event == 'dLimit':
            for x in values['dLimit'].split('/'):
                if x.isdigit():
                    Limit = int(x)
            window['dLimit1'].update(str(round(float(values['Length']) * 1000 / Limit, 2)) + ' mm')
            if float(values['Length']) * 1000 / Limit > d[0] + d[1]:
                window['dCheck'].update('OK')
            else:
                window['dCheck'].update('NG')
        if event == 'SectionType':
            if values['SectionType'] == 'Universal_Beam' or values['SectionType'] == 'Universal_Column' or values[
                'SectionType'] == 'PFC' or values['SectionType'] == 'RHS' or values['SectionType'] == 'SHS' or values[
                'SectionType'] == 'CHS':
                item = values[event]
                matrix = []
                SectionNames = wb[item]
                for x in SectionNames:
                    matrix.append(x[0].value)
                for i in range(7):
                    matrix.pop(0)
                SectionSize = list(filter(lambda item: item is not None, SectionSize))
                window['SectionSize'].update(value=matrix[7], values=matrix)
            elif values['SectionType'] == 'MGP 10' or values['SectionType'] == 'MGP 12':
                SectionSize = ['90x45','90x90','140x45', '140x90', '190x45','190x90', '240x45','240x90']
                window['SectionSize'].update(value='90x90', values=SectionSize)

        # window['b1'].update('OK')
        elif event == 'b2' or event == sg.WIN_CLOSED:
            break
        elif event == 'restraint':
            if values['restraint'] == 'FU' or values['restraint'] == 'PU':
                window['ends_with_restraint'].update(values=['Any'], value='Any')
            elif values['restraint'] == 'FF' or values['restraint'] == 'FP' or values['restraint'] == 'PP':
                window['ends_with_restraint'].update(values=['None', 'One', 'Both'], value='None')
            elif values['restraint'] == 'FL' or values['restraint'] == 'PL' or values['restraint'] == 'LL':
                window['ends_with_restraint'].update(values=['None'], value='None')

        elif event == 'print_calcs':
            SectionType1 = values['SectionType']
            SectionSize1 = values['SectionSize']
            Length = float(values['Length'])
            alpha_m = float(values['alpha_m'])
            restraint = values['restraint']
            load_height_position = values['load_height_position']
            longitudinal_position = values['longitudinal_position']
            ends_with_restraint = values['ends_with_restraint']
            section_properties = st.section_properties(SectionType1, SectionSize1, 0, 0, 0)
            calculations = st.compact(section_properties, SectionType1, Length, alpha_m, restraint,
                                      load_height_position,
                                      longitudinal_position, ends_with_restraint)
            st.printcalcs(SectionSize1, section_properties, values['print_location'], values['job_name'], '', 0, 0, 0,
                          True)
    window.close()




    # Create an event loop


variable = variables('Loading')
Layouts(variable)
